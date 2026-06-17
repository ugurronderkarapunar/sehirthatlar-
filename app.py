import pandas as pd
import numpy as np
from datetime import datetime
import os

# ------------------------------
# 1. VERİYİ OKU
# ------------------------------
file_path = '20260506_2_REV5.xlsx'

if not os.path.exists(file_path):
    print(f"❌ Hata: '{file_path}' dosyası bulunamadı. Lütfen dosyayı bu klasöre koyun.")
    exit()

print("📂 Dosya okunuyor...")
df = pd.read_excel(file_path, sheet_name='Sayfa2')

# Sütun adlarını temizle
df.columns = df.columns.str.strip()

# Tarih/saat dönüşümü (milisaniye sorununu çözer)
df['MERKEZDEN GEMİYE BİNİŞ'] = pd.to_datetime(
    df['MERKEZDEN GEMİYE BİNİŞ'], 
    format='mixed', 
    errors='coerce'
)
df = df.dropna(subset=['MERKEZDEN GEMİYE BİNİŞ'])

# Saat ve saat dilimi
df['Saat'] = df['MERKEZDEN GEMİYE BİNİŞ'].dt.hour
df['Saat_Dilimi'] = df['Saat'].apply(lambda x: f"{x:02d}:00–{x:02d}:59")

# Yön temizle
df['YÖN'] = df['YÖN'].str.strip()

print(f"✅ Veri yüklendi: {df.shape[0]} satır")

# ------------------------------
# 2. İKİ YÖN İÇİN AYRI DATAFRAME'LER
# ------------------------------
df_uskudar = df[df['YÖN'] == 'ÜSKÜDAR → BEŞİKTAŞ'].copy()
df_besiktas = df[df['YÖN'] == 'BEŞİKTAŞ → ÜSKÜDAR'].copy()

print(f"   Üsküdar→Beşiktaş: {len(df_uskudar)} yolcu")
print(f"   Beşiktaş→Üsküdar: {len(df_besiktas)} yolcu")

# ------------------------------
# 3. ANALİZ FONKSİYONU
# ------------------------------
def analiz_yap(df_yon, yon_adi):
    """
    Verilen yön için 4 temel analizi yapar:
    1. Saatlik biniş
    2. Geliş kaynakları (nereden geldi)
    3. Gidilen yerler (nereye gitti)
    4. Gitmeyenler (aktarma yapmayan)
    """
    
    # 1. Saat dilimine göre binme
    binme = df_yon.groupby('Saat_Dilimi').size().reset_index(name='Toplam_Binme')
    
    # 2. Geliş hatları (nereden geldi)
    gelis_df = df_yon[df_yon['MERKEZE GELDİĞİ ARACIN HATTI'].notna()]
    gelis_df = gelis_df[gelis_df['MERKEZE GELDİĞİ ARACIN HATTI'] != '']
    gelis = gelis_df.groupby(
        ['Saat_Dilimi', 'MERKEZE GELDİĞİ ARACIN HATTI']
    ).size().reset_index(name='Kisi_Sayisi')
    gelis.rename(columns={'MERKEZE GELDİĞİ ARACIN HATTI': 'Gelis_Hatti'}, inplace=True)
    
    # 3. İndikten sonra gidilen yerler (aktarma yapanlar)
    giden_df = df_yon[
        (df_yon['İNDİKTEN SONRA AKTARMA YAPTIMI(0=HAYIR,1=EVET)'] == 1) |
        (df_yon['İNDİKTEN SONRA NEREYE GİTTİ'].notna() & 
         df_yon['İNDİKTEN SONRA NEREYE GİTTİ'].str.strip().notna() & 
         (df_yon['İNDİKTEN SONRA NEREYE GİTTİ'] != 'BİLİNMİYOR'))
    ]
    gidis = giden_df.groupby(
        ['Saat_Dilimi', 'İNDİKTEN SONRA NEREYE GİTTİ']
    ).size().reset_index(name='Kisi_Sayisi')
    gidis.rename(columns={'İNDİKTEN SONRA NEREYE GİTTİ': 'Gidilen_Yer'}, inplace=True)
    
    # 4. İndikten sonra hiçbir yere gitmeyenler
    gitmeyen_df = df_yon[
        (df_yon['İNDİKTEN SONRA AKTARMA YAPTIMI(0=HAYIR,1=EVET)'] == 0) |
        (df_yon['İNDİKTEN SONRA NEREYE GİTTİ'].isna()) |
        (df_yon['İNDİKTEN SONRA NEREYE GİTTİ'].str.strip() == '') |
        (df_yon['İNDİKTEN SONRA NEREYE GİTTİ'] == 'BİLİNMİYOR')
    ]
    gitmeyen = gitmeyen_df.groupby('Saat_Dilimi').size().reset_index(name='Gitmeyen_Sayisi')
    
    return binme, gelis, gidis, gitmeyen

# ------------------------------
# 4. ANALİZLERİ ÇALIŞTIR
# ------------------------------
print("\n📊 Analizler yapılıyor...")
binme_u, gelis_u, gidis_u, gitmeyen_u = analiz_yap(df_uskudar, 'Üsküdar')
binme_b, gelis_b, gidis_b, gitmeyen_b = analiz_yap(df_besiktas, 'Beşiktaş')

# ------------------------------
# 5. TAHMİN VERİSİ HAZIRLA
# ------------------------------
print("🔮 Tahmin verisi hazırlanıyor...")

# Tüm gidilen yerleri topla
tum_gidis = df[
    (df['İNDİKTEN SONRA NEREYE GİTTİ'].notna()) &
    (df['İNDİKTEN SONRA NEREYE GİTTİ'].str.strip().notna()) &
    (df['İNDİKTEN SONRA NEREYE GİTTİ'] != 'BİLİNMİYOR')
].copy()
tum_gidis['Saat'] = tum_gidis['MERKEZDEN GEMİYE BİNİŞ'].dt.hour

# Yön + Hedef + Saat bazında ortalama
tahmin_df = tum_gidis.groupby(
    ['YÖN', 'İNDİKTEN SONRA NEREYE GİTTİ', 'Saat']
).size().reset_index(name='Tahmini_Yolcu_Sayisi')
tahmin_df.rename(columns={'İNDİKTEN SONRA NEREYE GİTTİ': 'Hedef'}, inplace=True)

# Tüm kombinasyonları oluştur (0-23 saat, tüm hedefler, her iki yön)
tum_hedefler = tahmin_df['Hedef'].unique()
tum_yonler = tahmin_df['YÖN'].unique()
tum_saatler = list(range(24))

tum_kombinasyon = []
for yon in tum_yonler:
    for hedef in tum_hedefler:
        for saat in tum_saatler:
            tum_kombinasyon.append({'YÖN': yon, 'Hedef': hedef, 'Saat': saat})

tum_kombinasyon_df = pd.DataFrame(tum_kombinasyon)

# Birleştir ve eksikleri 0 ile doldur
tahmin_full = tum_kombinasyon_df.merge(
    tahmin_df,
    on=['YÖN', 'Hedef', 'Saat'],
    how='left'
)
tahmin_full['Tahmini_Yolcu_Sayisi'] = tahmin_full['Tahmini_Yolcu_Sayisi'].fillna(0).astype(int)

# ------------------------------
# 6. PİVOT TABLOLARI HAZIRLA
# ------------------------------
# Üsküdar için pivot (Hedefler x Saat Dilimleri)
gidis_u_pivot = gidis_u.pivot(index='Gidilen_Yer', columns='Saat_Dilimi', values='Kisi_Sayisi').fillna(0)

# Beşiktaş için pivot
gidis_b_pivot = gidis_b.pivot(index='Gidilen_Yer', columns='Saat_Dilimi', values='Kisi_Sayisi').fillna(0)

# ------------------------------
# 7. ÖZET TABLO
# ------------------------------
ozet = pd.DataFrame({
    'Yön': ['Üsküdar → Beşiktaş', 'Beşiktaş → Üsküdar'],
    'Toplam_Biniş': [len(df_uskudar), len(df_besiktas)],
    'Aktarma_Yapan': [gidis_u['Kisi_Sayisi'].sum(), gidis_b['Kisi_Sayisi'].sum()],
    'Gitmeyen_(Aktarma_Yok)': [gitmeyen_u['Gitmeyen_Sayisi'].sum(), gitmeyen_b['Gitmeyen_Sayisi'].sum()],
    'Ortalama_Biniş/Saat': [round(len(df_uskudar)/24, 1), round(len(df_besiktas)/24, 1)]
})

# ------------------------------
# 8. EXCEL'E YAZ (openpyxl ile)
# ------------------------------
output_file = 'vapur_tam_analiz.xlsx'
print(f"\n📁 Excel dosyası oluşturuluyor: {output_file}")

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    
    # ---- ÜSKÜDAR YÖNÜ ----
    binme_u.to_excel(writer, sheet_name='Uskudar_Binme', index=False)
    gelis_u.to_excel(writer, sheet_name='Uskudar_Gelis', index=False)
    gidis_u.to_excel(writer, sheet_name='Uskudar_Gidis', index=False)
    gitmeyen_u.to_excel(writer, sheet_name='Uskudar_Gitmeyen', index=False)
    
    # ---- BEŞİKTAŞ YÖNÜ ----
    binme_b.to_excel(writer, sheet_name='Besiktas_Binme', index=False)
    gelis_b.to_excel(writer, sheet_name='Besiktas_Gelis', index=False)
    gidis_b.to_excel(writer, sheet_name='Besiktas_Gidis', index=False)
    gitmeyen_b.to_excel(writer, sheet_name='Besiktas_Gitmeyen', index=False)
    
    # ---- TAHMİN VERİSİ ----
    tahmin_full.to_excel(writer, sheet_name='Tahmin_Verisi', index=False)
    
    # ---- PİVOT TABLOLAR ----
    gidis_u_pivot.reset_index().to_excel(writer, sheet_name='Uskudar_Gidis_Pivot', index=False)
    gidis_b_pivot.reset_index().to_excel(writer, sheet_name='Besiktas_Gidis_Pivot', index=False)
    
    # ---- ÖZET ----
    ozet.to_excel(writer, sheet_name='Ozet', index=False)

# ------------------------------
# 9. EXCEL SÜTUN GENİŞLİKLERİNİ AYARLA (opsiyonel)
# ------------------------------
try:
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column].width = adjusted_width
    wb.save(output_file)
    print("✅ Sütun genişlikleri ayarlandı.")
except Exception as e:
    print(f"⚠️ Sütun genişliği ayarlanamadı: {e}")

# ------------------------------
# 10. SONUÇ RAPORU
# ------------------------------
print("\n" + "="*60)
print("📊 ANALİZ SONUÇLARI ÖZETİ")
print("="*60)

print(f"\n🚢 Üsküdar → Beşiktaş")
print(f"   Toplam Yolcu: {len(df_uskudar)}")
print(f"   Aktarma Yapan: {gidis_u['Kisi_Sayisi'].sum()}")
print(f"   Gitmeyen: {gitmeyen_u['Gitmeyen_Sayisi'].sum()}")
print(f"   En Yoğun Saat: {binme_u.loc[binme_u['Toplam_Binme'].idxmax(), 'Saat_Dilimi']} ({binme_u['Toplam_Binme'].max()} kişi)")
print(f"   En Çok Gidilen Yer: {gidis_u.groupby('Gidilen_Yer')['Kisi_Sayisi'].sum().idxmax()}")

print(f"\n🚢 Beşiktaş → Üsküdar")
print(f"   Toplam Yolcu: {len(df_besiktas)}")
print(f"   Aktarma Yapan: {gidis_b['Kisi_Sayisi'].sum()}")
print(f"   Gitmeyen: {gitmeyen_b['Gitmeyen_Sayisi'].sum()}")
print(f"   En Yoğun Saat: {binme_b.loc[binme_b['Toplam_Binme'].idxmax(), 'Saat_Dilimi']} ({binme_b['Toplam_Binme'].max()} kişi)")
print(f"   En Çok Gidilen Yer: {gidis_b.groupby('Gidilen_Yer')['Kisi_Sayisi'].sum().idxmax()}")

print("\n" + "="*60)
print(f"✅ Tüm analizler '{output_file}' dosyasına kaydedildi.")
print("📌 Müdürün sorusu için: 'Uskudar_Gidis_Pivot' veya 'Besiktas_Gidis_Pivot' sayfalarını açın.")
print("   Örneğin: Beykoz'dan gelenleri 'Uskudar_Gelis' sayfasında, Sarıyer'e gidenleri 'Uskudar_Gidis_Pivot'da filtreleyin.")
print("="*60)
